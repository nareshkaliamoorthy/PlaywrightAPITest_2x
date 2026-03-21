from openpyxl import load_workbook

def get_xlsx_data (filepath, sheet="api_payloads_v2"):
    with open (filepath) as file:
        data = []
        workbook = load_workbook(filepath,sheet)
        sheet = workbook[sheet]

        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data)

    return data
